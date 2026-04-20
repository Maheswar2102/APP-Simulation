"""Excel-based runtime configuration loader."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class RuntimeConfig:
    app_username: str
    app_password: str
    customer_name: str | None = None
    configuration_number: str | None = None
    menu_group: str | None = None
    menu_item: str | None = None


@dataclass
class HierarchyLevelConfig:
    name: str
    visible: bool = True
    non_hierarchial: bool = False


@dataclass
class HierarchyColumnConfig:
    hierarchy_key: str
    hierarchy_name: str
    levels: list[HierarchyLevelConfig]


@dataclass
class CrossHierarchyConfig:
    attribute_name: str
    hierarchy_1_product: str
    hierarchy_2_channel: str
    hierarchy_3_location: str
    attribute_type: str | None = None
    mapped_column: str | None = None
    editable: str | None = None


def load_menu_target_from_sheet(file_path: str, sheet_name: str) -> str:
    """Return the first non-empty cell text from the given sheet (used as menu_item)."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(f"Only Excel files are supported (.xlsx/.xlsm). Got: {path.name}")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            value = ws.cell(row=r, column=c).value
            text = "" if value is None else str(value).strip()
            if text:
                return text

    raise ValueError(f"Sheet '{sheet_name}' has no non-empty cells to use as menu target")


def load_sheet_rows_as_dicts(file_path: str, sheet_name: str) -> list[dict[str, str]]:
    """
    Read a sheet as row dictionaries using header row (row 1).
    Returns only rows that contain at least one non-empty cell.
    Header keys are normalized to lowercase with spaces.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(f"Only Excel files are supported (.xlsx/.xlsm). Got: {path.name}")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]

    if ws.max_row < 2:
        return []

    headers: list[str] = []
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=c).value
        h = str(raw or "").strip().lower()
        headers.append(h)

    rows: list[dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        row_dict: dict[str, str] = {}
        has_any = False
        for c in range(1, ws.max_column + 1):
            key = headers[c - 1]
            if not key:
                continue
            val = ws.cell(row=r, column=c).value
            text = "" if val is None else str(val).strip()
            if text:
                has_any = True
            row_dict[key] = text
        if has_any:
            rows.append(row_dict)

    return rows


def load_named_table_rows_from_sheet(
    file_path: str,
    sheet_name: str,
    header_first_col: str,
) -> list[dict[str, str]]:
    """
    Read a table whose header row can appear anywhere in the sheet.

    The header row is identified by the first-column header text (case-insensitive),
    then subsequent non-empty rows are returned as dictionaries until a fully blank row.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(f"Only Excel files are supported (.xlsx/.xlsm). Got: {path.name}")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]

    header_row_idx: int | None = None
    target = header_first_col.strip().lower()
    for r in range(1, ws.max_row + 1):
        first_val = ws.cell(row=r, column=1).value
        first_text = "" if first_val is None else str(first_val).strip().lower()
        if first_text == target:
            header_row_idx = r
            break

    if header_row_idx is None:
        return []

    headers: list[str] = []
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row_idx, column=c).value
        h = str(raw or "").strip()
        headers.append(h)

    # Use only non-empty headers and preserve order.
    active_cols: list[tuple[int, str]] = []
    for idx, h in enumerate(headers, start=1):
        if h:
            active_cols.append((idx, h))

    if not active_cols:
        return []

    rows: list[dict[str, str]] = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_dict: dict[str, str] = {}
        has_any = False
        for c, key in active_cols:
            val = ws.cell(row=r, column=c).value
            text = "" if val is None else str(val).strip()
            if text:
                has_any = True
            row_dict[key] = text

        if not has_any:
            break
        rows.append(row_dict)

    return rows


def load_hierarchy_configs_from_master_data(file_path: str, sheet_name: str = "Master data") -> list[HierarchyColumnConfig]:
    """
    Parse hierarchy config from the Master data sheet in row-wise block format.

    Each hierarchy block:
      - A row with "Hierarchy N" in col1 (col2/col3 empty) = block start
      - A row with the hierarchy name in col1 (col2/col3 empty OR column header labels)
      - Optional explicit column header row: Name | visible | Non-Hierarchial
      - Level rows: col1=level name, col2=visible (yes/no), col3=Non-Hierarchial (yes/no)
    """
    import re as _re

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(f"Only Excel files are supported (.xlsx/.xlsm). Got: {path.name}")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]

    if ws.max_row < 3:
        return []

    # Find all rows that start a hierarchy block ("Hierarchy N" with no other columns)
    hierarchy_key_rows: list[int] = []
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        c3 = ws.cell(row=r, column=3).value
        text = "" if c1 is None else str(c1).strip()
        if _re.match(r"hierarchy\s*\d+", text, _re.IGNORECASE) and not c2 and not c3:
            hierarchy_key_rows.append(r)

    configs: list[HierarchyColumnConfig] = []
    for idx, start_row in enumerate(hierarchy_key_rows):
        key = str(ws.cell(row=start_row, column=1).value).strip()
        end_row = hierarchy_key_rows[idx + 1] - 1 if idx + 1 < len(hierarchy_key_rows) else ws.max_row

        hierarchy_name: str | None = None
        levels: list[HierarchyLevelConfig] = []

        for r in range(start_row + 1, end_row + 1):
            v1 = "" if ws.cell(row=r, column=1).value is None else str(ws.cell(row=r, column=1).value).strip()
            v2 = "" if ws.cell(row=r, column=2).value is None else str(ws.cell(row=r, column=2).value).strip()
            v3 = "" if ws.cell(row=r, column=3).value is None else str(ws.cell(row=r, column=3).value).strip()

            if not v1:
                continue

            v2l = v2.lower()
            v3l = v3.lower()

            # Column header row: col2 or col3 contain "visible" or "non-hierarchi"
            if "visible" in v2l or "non-hierarchi" in v2l or "visible" in v3l or "non-hierarchi" in v3l:
                # col1 may also carry the hierarchy name if it's not a generic label
                if v1.lower() != "name":
                    hierarchy_name = hierarchy_name or v1
                continue

            # Pure hierarchy name row (only col1 filled, not a generic label)
            if not v2 and not v3 and v1.lower() != "name":
                hierarchy_name = hierarchy_name or v1
                continue

            # Level data row: col2/col3 are yes/no
            if v2l in ("yes", "no") or v3l in ("yes", "no"):
                levels.append(HierarchyLevelConfig(
                    name=v1,
                    visible=(v2l == "yes"),
                    non_hierarchial=(v3l == "yes"),
                ))

        if hierarchy_name and levels:
            configs.append(HierarchyColumnConfig(
                hierarchy_key=key,
                hierarchy_name=hierarchy_name,
                levels=levels,
            ))

    return configs


def load_cross_hierarchy_configs_from_master_data(file_path: str, sheet_name: str = "Master data") -> list[CrossHierarchyConfig]:
    """
    Parse cross hierarchy config from the Master data sheet.

    Structure:
      - A row with "Cross Hierarchy" in col1 (col2/col3 empty) = section start marker
      - Header row: Attribute Name | Hierarchy 1\nProduct | Hierarchy 2 Channel | Hierarchy Location | Editable | Cross Hierarchy
      - Data rows: each row has attribute name and hierarchy mappings
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(f"Only Excel files are supported (.xlsx/.xlsm). Got: {path.name}")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]

    if ws.max_row < 49:
        return []

    # Find "Cross Hierarchy" marker row
    cross_h_marker_row: int | None = None
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value
        text = "" if c1 is None else str(c1).strip()
        if text.lower() == "cross hierarchy":
            cross_h_marker_row = r
            break

    if cross_h_marker_row is None:
        return []

    def _nh(v: Any) -> str:
        return str(v or "").strip().lower().replace(" ", "").replace("_", "").replace("\n", "")

    def _detect_header_indices(row_values: list[str]) -> tuple[int, int, int, int, int | None, int | None, int | None] | None:
        attr_name_col = None
        h1_col = None
        h2_col = None
        h3_col = None
        attr_type_col = None
        mapped_col = None
        editable_col = None

        for idx, raw in enumerate(row_values):
            h = _nh(raw)
            if "attribute" in h and "name" in h:
                attr_name_col = idx
            elif "product" in h and "hierarchy" in h:
                h1_col = idx
            elif "channel" in h and "hierarchy" in h:
                h2_col = idx
            elif "location" in h and "hierarchy" in h:
                h3_col = idx
            elif "attribute" in h and "type" in h:
                attr_type_col = idx
            elif "mapped" in h and "column" in h:
                mapped_col = idx
            elif "map" in h and "to" in h:
                mapped_col = idx
            elif "editable" in h:
                editable_col = idx

        if attr_name_col is None or h1_col is None or h2_col is None or h3_col is None:
            return None
        return (attr_name_col, h1_col, h2_col, h3_col, attr_type_col, mapped_col, editable_col)

    def _cell_str(row_idx: int, col_idx: int | None) -> str:
        if col_idx is None:
            return ""
        v = ws.cell(row=row_idx, column=col_idx + 1).value
        return "" if v is None else str(v).strip()

    current_cols: tuple[int, int, int, int, int | None, int | None, int | None] | None = None
    configs: list[CrossHierarchyConfig] = []

    # Parse all data blocks under "Cross Hierarchy" (set1, set2, ...), even when separated by blank rows.
    for r in range(cross_h_marker_row + 1, ws.max_row + 1):
        row_values = [str(ws.cell(row=r, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]

        detected = _detect_header_indices(row_values)
        if detected is not None:
            current_cols = detected
            continue

        if current_cols is None:
            continue

        attr_name_col, h1_col, h2_col, h3_col, attr_type_col, mapped_col, editable_col = current_cols
        attr_name = _cell_str(r, attr_name_col)
        if not attr_name:
            continue

        attr_norm = _nh(attr_name)
        # Ignore block labels and any accidental repeated header text in data rows.
        if attr_norm.startswith("set"):
            continue
        if "attribute" in attr_norm and "name" in attr_norm:
            continue

        h1 = _cell_str(r, h1_col)
        h2 = _cell_str(r, h2_col)
        h3 = _cell_str(r, h3_col)
        if not (h1 or h2 or h3):
            continue

        attribute_type = _cell_str(r, attr_type_col)
        mapped_column = _cell_str(r, mapped_col)
        editable = _cell_str(r, editable_col)

        configs.append(CrossHierarchyConfig(
            attribute_name=attr_name,
            hierarchy_1_product=h1,
            hierarchy_2_channel=h2,
            hierarchy_3_location=h3,
            attribute_type=attribute_type,
            mapped_column=mapped_column,
            editable=editable,
        ))

    return configs


def _normalize_header(name: Any) -> str:
    return str(name or "").strip().lower().replace(" ", "").replace("_", "")


ALIASES = {
    "app_username": ["appusername", "username", "userid", "user"],
    "app_password": ["apppassword", "password", "passwd", "pwd"],
    "customer_name": ["customername", "customer", "custname"],
    "config_number": ["confignumber", "configurationnumber", "configurationno", "configno", "configuration"],
    "menu_group": ["menugroup", "leftmenu", "dropdown", "module", "section", "sheetname"],
    "menu_item": ["menuitem", "submenu", "option", "item", "target", "childmenu"],
}


def _resolve_header_map(raw_headers: list[Any]) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, value in enumerate(raw_headers):
        key = _normalize_header(value)
        if key:
            headers[key] = idx
    return headers


def _find_col_idx(headers: dict[str, int], canonical: str) -> int:
    for alias in ALIASES[canonical]:
        if alias in headers:
            return headers[alias]
    raise ValueError(f"Missing required column for {canonical}. Found headers: {sorted(headers.keys())}")


def _find_optional_col_idx(headers: dict[str, int], canonical: str) -> int | None:
    for alias in ALIASES[canonical]:
        if alias in headers:
            return headers[alias]
    return None


def _read_required(value: Any, field_name: str, row_number: int) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        raise ValueError(f"Empty value for {field_name} at row {row_number}")
    return text


def _read_optional(value: Any) -> str | None:
    text = "" if value is None else str(value).strip()
    return text or None


def load_runtime_config_from_excel(file_path: str, sheet_name: str | None = None, row_index: int = 1) -> RuntimeConfig:
    """
    Read runtime fields from an Excel sheet.

    Expected headers (case-insensitive, spaces/underscores ignored):
    - app_username
    - app_password
    - customer_name
    - config_number
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(
            f"Only Excel files are supported (.xlsx/.xlsm). Got: {path.name}"
        )

    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    raw_headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    headers = _resolve_header_map(raw_headers)
    data_row = row_index + 1  # row_index=1 means first data row under header

    app_username = _read_required(
        ws.cell(row=data_row, column=_find_col_idx(headers, "app_username") + 1).value,
        "app_username",
        data_row,
    )
    app_password = _read_required(
        ws.cell(row=data_row, column=_find_col_idx(headers, "app_password") + 1).value,
        "app_password",
        data_row,
    )

    cust_col = _find_optional_col_idx(headers, "customer_name")
    conf_col = _find_optional_col_idx(headers, "config_number")
    menu_group_col = _find_optional_col_idx(headers, "menu_group")
    menu_item_col = _find_optional_col_idx(headers, "menu_item")

    customer_name = _read_optional(ws.cell(row=data_row, column=cust_col + 1).value if cust_col is not None else None)
    configuration_number = _read_optional(ws.cell(row=data_row, column=conf_col + 1).value if conf_col is not None else None)
    menu_group = _read_optional(ws.cell(row=data_row, column=menu_group_col + 1).value if menu_group_col is not None else None)
    menu_item = _read_optional(ws.cell(row=data_row, column=menu_item_col + 1).value if menu_item_col is not None else None)

    # Fallback for XLSX: sheet name can act as top-level menu group (e.g., "Master data")
    if not menu_group and sheet_name:
        menu_group = sheet_name

    # If menu_item column is not present, try first non-empty non-credential cell in data row
    if not menu_item:
        used_cols = {
            _find_col_idx(headers, "app_username"),
            _find_col_idx(headers, "app_password"),
        }
        if cust_col is not None:
            used_cols.add(cust_col)
        if conf_col is not None:
            used_cols.add(conf_col)
        if menu_group_col is not None:
            used_cols.add(menu_group_col)

        for col in range(ws.max_column):
            if col in used_cols:
                continue
            candidate = _read_optional(ws.cell(row=data_row, column=col + 1).value)
            if candidate:
                menu_item = candidate
                break

    return RuntimeConfig(
        app_username=app_username,
        app_password=app_password,
        customer_name=customer_name,
        configuration_number=configuration_number,
        menu_group=menu_group,
        menu_item=menu_item,
    )
